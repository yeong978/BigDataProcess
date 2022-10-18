#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb["Sheet1"]


row_id = 1;
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value
        ws.cell(row = row_id, column = 7).value = sum_v
    row_id += 1


row_id = 1
sum_v = 0
for row in ws:
    if row_id != 1:       
        sum_v += ws.cell(row = row_id, column = 7).value        
    row_id += 1   

middle = round(sum_v/(row_id-2), 2) 

a_m = middle + round(middle*0.4*0.5,2) 
a_p = a_m + round(a_m*0.3*0.5,2) 
b_m = middle
c_m = middle - round(middle*0.4*0.5,2) 
c_p = c_m - round(middle*0.3*0.5,2) 

row_id1 = 1

for row in ws:   
    if row_id1 != 1:        
        if ws.cell(row = row_id1, column = 7).value > a_m:
            if ws.cell(row = row_id1, column = 7).value > a_p:
                ws.cell(row = row_id1, column = 8).value = 'A+'
            else:
                ws.cell(row = row_id1, column = 8).value = 'A0'            
        elif ws.cell(row = row_id1, column = 7).value < c_m:
            if ws.cell(row = row_id1, column = 7).value > c_p:
                ws.cell(row = row_id1, column = 8).value = 'C+'
            else:
                ws.cell(row = row_id1, column = 8).value = 'C0'            
        else:
            if ws.cell(row = row_id1, column = 7).value > b_m:
                ws.cell(row = row_id1, column = 8).value = 'B+'
            else:
                ws.cell(row = row_id1, column = 8).value = 'B0'
    row_id1 += 1
    
wb.save("student.xlsx")

